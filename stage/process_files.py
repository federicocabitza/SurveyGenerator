from Group import Group
from Answer import Answer
from Survey import Survey
from RecordExcel import RecordExcel
from SurveyLanguage import SurveyLanguage
from Survey_Question import Survey_Question
from Question_Attributes import Question_Attributes

import sys
import os
import re
import codecs
import random
import datetime
import openpyxl as xl
import xml.etree.ElementTree as ET
from deep_translator import GoogleTranslator

def process_files( admin_email, admin_name, survey_id, primary_language, survey_title, survey_description, survey_welcometext,
                  survey_endtext, survey_alias, iterate_group, *args ):

    dir_current = os.path.dirname( os.path.abspath(__file__) ) #current directory
    dir_config = os.path.join( dir_current, 'config' ) #config directory
    filepaths = []
    secondary_languages = ""
    xlsxFile = None
    survey_id = random.randint(10000, 999999) if survey_id == None else survey_id
    iterate_group = 0 if iterate_group == None else iterate_group
    filepaths, xlsxFile, secondary_languages = populateFiles_andLangs( *args )

    today_date = datetime.datetime.now().strftime("%Y-%m-%d")
    output_filename = f"Survey_{today_date}.txt"
    file_header = os.path.join( dir_config, "header.txt" )

    config_survey = Survey( survey_id, admin_email, primary_language, admin_name, secondary_languages)

    header = read_header( file_header )
    write = codecs.open( output_filename, "w", "utf-8" )
    write.write(header + '\n')
    write.write(config_survey.__str__())
    write.write( get_languages_structure( survey_id, primary_language, survey_title, survey_description,
       secondary_languages, survey_welcometext, survey_endtext, survey_alias ))


    count = 0
    if xlsxFile != None:
        list_records = readXLSX( xlsxFile )
        if list_records[0].id != None:
            for i in range( iterate_group ):
                for record in list_records:
                    for filepath in filepaths:
                        parse_toTXT( write, filepath, count, list_records, record)
                        count += 1
        else:
            for i in range( iterate_group ):
                for filepath in filepaths:
                    parse_toTXT( write, filepath, count, list_records, None)
                    count += 1


    else:
        for i in range( iterate_group ):
            for filepath in filepaths:
                parse_toTXT( write, filepath, count, None, None)
                count += 1


#METHODS#
def parse_toTXT( write, file, i, list_records=None, record=None):
    tree = ET.parse( file )
    root = tree.getroot()
    dictionary_comprehension = { child: parent for parent in root.iter() for child in parent }

    qidSommaFine = ( len(list_records) - 1) * 2 if list_records != None else 2
    typeSommaFine = ( len(list_records) - 1) * 3 if list_records != None else 3
    list_group = []
    list_group_l10ns = []
    list_question = []
    list_question_l10ns = []
    list_answer = []
    list_answer_l10ns = []
    list_subquestion = []
    list_questionAttr = []

    for row in root.iter("row"):
        collection_attrs = {}
        for child in row:
            tag = child.tag
            if( tag in ["type", "encrypted", "id", "class"] ):
                collection_attrs[ "_" + tag ] = child.text
            else:
                collection_attrs[ child.tag ] = child.text

        if list_records is None:
            collection_attrs = replace_tags( collection_attrs, primary_language, None)
        else:
            collection_attrs = replace_tags( collection_attrs, list_records, primary_language )

        condition = dictionary_comprehension[ dictionary_comprehension[ row ] ].tag

        match condition:

            case "groups" | "group_l10ns":

                temp_group = Group( **collection_attrs )
                temp_group.gid = str( int(temp_group.gid) + qidSommaFine + (3 * i) )
                temp_group.group_order = str( int(temp_group.group_order) + typeSommaFine + (3 * i) )

                list_group.append( temp_group ) if condition == 'groups' else list_group_l10ns.append( temp_group )

            case "questions" | "question_l10ns" | "subquestions":

                temp_question = Survey_Question( **collection_attrs )
                temp_question.qid = str( int(temp_question.qid) + 10 * i) if temp_question.qid != None else None
                temp_question.gid = str( int(temp_question.gid) + qidSommaFine + (3 * i) ) if temp_question.gid != None else None

                if condition != "subquestions":
                    temp_question.title = temp_question.setTitle(i)
                else:
                    temp_question.parent_qid = str( int(temp_question.parent_qid) + 10 * i)

                if temp_question.question != None and "<URL>" in temp_question.question and record != None:
                    temp_question.question = temp_question.question.replace("<URL>", "https://www.entechne.com/ecg/{}.gif".format(str(record.id)))

                if temp_question.question != None and "ADVICE" in temp_question.question and record != None:
                    temp_question.question = temp_question.question.replace("ADVICE", str(record.advice)).replace("<INFORMAZIONI-CONTESTO>", str(record.contesto))

                temp_question._class = "SQ" if condition == 'subquestions' else "Q"
                temp_question._type = None if condition == 'subquestions' else temp_question._type
                if condition == 'subquestions':
                    list_subquestion.append( temp_question )
                elif condition == 'questions':
                    list_question.append( temp_question )
                else:
                    list_question_l10ns.append( temp_question )

            case "answers" | "answer_l10ns":
                temp_answer = Answer( **collection_attrs )

                temp_answer.qid = str( int(temp_answer.qid) + 10 * i ) if temp_answer.qid != None else None
                temp_answer.answer = "\"{}\"".format( temp_answer.answer ) if temp_answer.answer != None else None

                if temp_answer.answer == "No, non accetto e scelgo la categoria opposta.":
                    temp_answer.assessment_value = record.assessmentA[0] if record != None else None

                if temp_answer.answer == "Sì, accetto la categoria proposta dalla macchina":
                    temp_answer.assessment_value = record.assessmentA[1] if record != None else None

                list_answer.append( temp_answer ) if condition == 'answers' else list_answer_l10ns.append( temp_answer )

            case "question_attributes":
                temp_question_attributes = Question_Attributes( **collection_attrs )
                temp_question_attributes.qid = str( int(temp_question_attributes.qid) + 10 * i ) if temp_question_attributes.qid != None else None
                list_questionAttr.append( temp_question_attributes )

            case _:
                print( condition )


    if len(list_questionAttr) != 0 and len(list_question_l10ns) == 0:
        list_question = set_attributes( list_question, list_questionAttr )

    if len( list_group_l10ns) != 0:
        g_informations = []
        gq_informations = []
        gsq_informations = []
        ga_informations = []

        #Group added
        for group in list_group:
            for group_l10n in list_group_l10ns:
                if group.gid == group_l10n.gid:
                    g_informations.append( group_l10n | group )

        #Questions
        for question in list_question:
            for quest10n in list_question_l10ns:
                for group_l10n in g_informations:
                    if question.qid == quest10n.qid:
                        if question.gid == group_l10n.gid and quest10n.language == group_l10n.language:
                            join_question = quest10n | question
                            gq_informations.append( join_question )

        print(len(gq_informations))

        if len(list_questionAttr) != 0 and len(gq_informations) != 0:
            print(len(gq_informations))
            gq_informations = set_attributes( gq_informations, list_questionAttr )
            print(len(gq_informations))

        #Subquestions
        for subquestion in list_subquestion:
            for quest10n in list_question_l10ns:
                for group_l10n in list_group_l10ns:
                    if question.qid == subquestion.parent_qid and quest10n.qid == subquestion.qid and quest10n.language == group_l10n.language:
                        join_subquestion = quest10n | subquestion
                        join_subquestion._class = 'SQ'
                        gsq_informations.append( join_subquestion )

        #Answers
        for question in gq_informations:
            for answer in list_answer:
                for answer_l10n in list_answer_l10ns:
                    if question.qid == answer.qid and answer.aid == answer_l10n.aid and question.language == answer_l10n.language:
                        join_answer = answer | answer_l10n
                        ga_informations.append( join_answer )

        write_file( write, g_informations, gq_informations, gsq_informations, ga_informations)

    else:
        write_file( write, list_group, list_question, list_subquestion, list_answer, record)

def write_file( write, groups, questions, subquestions, answers, record=None):
    for group in groups:
        write.write( group.__str__() + "\n" )
        for question in questions:
            if question.gid == group.gid and question.language == group.language:
                write.write( question.__str__() + "\n" )

                for subquestion in subquestions:
                    if question.qid == subquestion.parent_qid and question.language == subquestion.language:
                        write.write( subquestion.__str__() + "\n" )

                for answer in answers:
                    if question.qid == answer.qid and question.language == answer.language:
                        answer.assessment_value = set_assessment_value( record, question.title, answer ) if record != None else answer.assessment_value
                        write.write( answer.__str__() + "\n" )


def set_assessment_value( record, question_title, answer ):
    if question_title == str(record.id) + "Caso":
        if answer == "Assenza di fratture" and (record.assessment == 0 or record.assessment == 1):
            return 0
        if answer == "Presenza di fratture" and (record.assessment == 0 or record.assessment == 1):
            return 1


def set_attributes( list_question, list_attrs ):
    for question in list_question:
        for attribute in list_attrs:
            if question.qid == attribute.qid and hasattr( question, attribute.attribute ):
                setattr( question, attribute.attribute, attribute.value )
    return list_question


def replace_tags( collection, primary_language, records=None ):
    language = collection.get('language')
    if not language:
        return collection
    if records is None:
        return collection

    for key, value in collection.items():
        if value is not None:
            for record in records:
                pattern = re.compile( rf'\[{record.etichetta}([A-Z])*\]')
                if pattern.search( collection[key] ):
                    if language != primary_language:
                        translator = GoogleTranslator( source= primary_language, target=language )
                        translated_text = translator.translate( record.esempio )
                    else:
                        translated_text = record.esempio

                    collection[key] = pattern.sub( translated_text, collection[key])
    return collection


def populateFiles_andLangs( *args ):
    filepaths = []
    xlsx_file = None
    secondary_languages = ""
    for arg in args:
        if os.path.exists(arg):
            if not arg.endswith('.xlsx'):
                filepaths.append(arg)
            else:
                xlsx_file = arg
        elif arg != primary_language:
            secondary_languages += arg + " "

    return filepaths, xlsx_file, secondary_languages


def readXLSX( xlsx_file ):
    workbook = xl.load_workbook( xlsx_file )
    worksheet = workbook.worksheets[0] #get the excel's first sheet

    indexes = []
    list_records = []
    for index in worksheet[1]: #With this loop i take the indexes present in the excel file (1st row)
        if index.value is None:
            break
        indexes.append( index.value )

    for row in worksheet.iter_rows( 2, worksheet.max_row ):
        excel_data = {}
        i = 0

        #if the current column doesn't have any value, then breaks the loop
        if all( cell.value is None for cell in row ):
            break

        for index in indexes:
            tag = index.lower() if index not in ["assessmentA", "assessmentR"] else index
            #if row doesn't start with '='. I add to the variable, else i do nothing
            if row[i].value and not str( row[i].value ).startswith('='):
                excel_data[tag] = row[i].value
            else:
                excel_data[tag] = None
            i += 1

        record = RecordExcel( **excel_data )
        list_records.append( record )

    return list_records


#This method allows to translate some phrases written by the user from the primary_language to secondary_languages
def get_languages_structure( survey_id, primary_language, survey_title, survey_description,
   secondary_languages=None, survey_welcometext=None, survey_endtext=None, survey_alias=None ):

    conf_firstLang = SurveyLanguage( survey_id, primary_language, survey_title,
       survey_description, survey_welcometext, survey_endtext, survey_alias)
    structure_str = conf_firstLang.__str__()

    if secondary_languages is not None and secondary_languages != "":
        secondary_languages = secondary_languages[:-1]
        languages = secondary_languages.split(' ')

        for current_language in languages:
            translator = GoogleTranslator(source=primary_language, target=current_language)

            trans_title = translator.translate( survey_title )
            trans_desc = translator.translate( survey_description )
            trans_wc = translator.translate( survey_welcometext ) if survey_welcometext is not None else None
            trans_end = translator.translate( survey_endtext ) if survey_endtext is not None else None
            trans_alias = translator.translate( survey_alias ) if survey_alias is not None else None

            conf_otherLang = SurveyLanguage( survey_id, current_language, trans_title,
            trans_desc, trans_wc, trans_end, trans_alias )

            structure_str += conf_otherLang.__str__()

    return structure_str


def read_header( filepath ):
    with codecs.open( filepath, 'r', 'utf-8' ) as file:
        lines = file.read().splitlines()
    header = '\t'.join( lines )
    return header


if __name__ == "__main__":
    admin_email = sys.argv[1]
    admin_name = sys.argv[2]
    survey_id_str = sys.argv[3]
    if survey_id_str:
        try:
            survey_id = int( survey_id_str )
        except ValueError:
            print( f"Errore: survey_id '{survey_id_str}' non è un numero valido." )
            survey_id = None
    else:
        survey_id = None

    iterate_group_str = sys.argv[10]
    if iterate_group_str:
        try:
            iterate_group = int( iterate_group_str )
        except ValueError:
            print( f"Errore: iterate_group '{iterate_group_str}' non è un numero valido." )
            iterate_group = 0
    else:
        iterate_group = None
    primary_language = sys.argv[4]
    survey_title = sys.argv[5]
    survey_description = sys.argv[6]

    #Parameters - Optional
    survey_welcometext = sys.argv[7] if sys.argv[7] is not None else ""
    survey_endtext = sys.argv[8] if sys.argv[8] is not None else ""
    survey_alias = sys.argv[9] if sys.argv[9] is not None else ""

    process_files( admin_email, admin_name, survey_id, primary_language, survey_title,
        survey_description, survey_welcometext, survey_endtext, survey_alias, iterate_group, *sys.argv[11:] )
